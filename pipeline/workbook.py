"""Phase 4 workbook: shows the work, not just the answers.

Intermediate columns stay visible and the composite value, VORP, ROI and tier
cells are REAL EXCEL FORMULAS reading from weight cells on Assumptions, so the
reader can change a weight and watch the ranking move.
"""
import pandas as pd, numpy as np, json
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, LineChart, ScatterChart, Reference, Series

P = '/home/claude/p4/'
OUT = '/mnt/user-data/outputs/emerald_imperium/Phase4_Valuation.xlsx'

VAL = pd.read_pickle(P + 'VAL.pkl')
CPT = pd.read_pickle(P + 'CPT.pkl')
VALID = pd.read_pickle(P + 'VALID.pkl')
V = pd.read_pickle(P + 'valuation.pkl')

F = 'Arial'
HDR = Font(name=F, bold=True, color='FFFFFF', size=10)
HFILL = PatternFill('solid', fgColor='2F4F6F')
BLUE = Font(name=F, color='0000FF', size=10)      # hardcoded input / lever
BLACK = Font(name=F, size=10)                      # formula
GREEN = Font(name=F, color='008000', size=10)      # cross-sheet link
YELL = PatternFill('solid', fgColor='FFFF00')
BOLD = Font(name=F, bold=True, size=10)

wb = Workbook()

# ============================ README ==================================
ws = wb.active; ws.title = 'README'
ws.column_dimensions['A'].width = 34
ws.column_dimensions['B'].width = 110
rows = [
    ('Pokémon Emerald Imperium — Phase 4 Valuation', ''),
    ('', ''),
    ('Parse date', '2026-09-01'),
    ('Run settings', 'NORMAL difficulty; Minimal Grinding Mode OFF; level caps enforced; '
                     'standard playthrough, no breeding (RUN_CONFIG.md)'),
    ('Damage engine', 'RadicalRedShowdown/damage-calc @ 7f35400, @smogon/calc/adaptable, '
                      'driven by a custom Imperium data layer'),
    ('Level cap ladder', 'MEASURED from Boss_Battles.xlsx!Main rows 9-27, supplied 2026-09-01. '
                     'Supersedes the earlier reconstruction, which was right on 18 of 19 caps '
                     '(cp09 was interpolated at 53; it is 56).'),
    ('Generation pinned', '9 — Inferred. Evidence: Tera Blast/Ivy Cudgel/Armor Cannon present; '
                          'all 18 types carry both physical and special moves (per-move split)'),
    ('Engine verification', '14/14 damage calcs match an independent hand implementation of the '
                            'formula. Discrepancy rate 0.0%'),
    ('', ''),
    ('Source files', 'rows'),
    ('01_species.tsv (Phase 1)', 1534),
    ('02_world.tsv (Phase 2)', 419),
    ('03_trainers.tsv (Phase 3)', 164),
    ('03_trainer_slots.tsv (Phase 3)', 678),
    ('imperium_layer.json', '1,534 species / 937 moves / 18-type chart / 327 abilities'),
    ('', ''),
    ('Rows scored', '42,026 = 1,311 obtainable forms x 19 checkpoints x {floor, ceiling}'),
    ('Scoring errors', 0),
    ('', ''),
    ('DATA INTEGRITY LOG', 'see the Integrity sheet'),
    ('Colour key', 'BLUE = hardcoded input you may change  |  BLACK = formula  |  '
                   'GREEN = link to another sheet  |  YELLOW = key assumption'),
]
for i, (a, bb) in enumerate(rows, 1):
    ws.cell(i, 1, a).font = BOLD if (i == 1 or a.isupper() or a in
                                     ('Source files', 'Colour key')) else Font(name=F, size=10)
    ws.cell(i, 2, bb).font = Font(name=F, size=10)
    ws.cell(i, 2).alignment = Alignment(wrap_text=True, vertical='top')

# ========================= ASSUMPTIONS ================================
aw = wb.create_sheet('Assumptions')
aw.column_dimensions['A'].width = 30
aw.column_dimensions['B'].width = 12
aw.column_dimensions['C'].width = 92
aw.cell(1, 1, 'COMPOSITE VALUE WEIGHTS — change these and the whole workbook recalculates').font = BOLD
hdr = ['Component', 'Weight', 'Why this weight']
for j, h in enumerate(hdr, 1):
    c = aw.cell(2, j, h); c.font = HDR; c.fill = HFILL
W = [('w_offense', 0.26, 'Computed from actual damage rolls; the most load-bearing axis'),
     ('w_defense', 0.15, 'Single-hit survival, also from damage rolls'),
     ('w_sustain', 0.16, 'Sweep depth on one lifebar with recovery, screens and hazard chip, '
                         'minus the same species with no kit. 0.24 for defensive roles'),
     ('w_speed', 0.12, 'Roster outsped. 0.04 for defensive roles, which do not want the stat'),
     ('w_utility', 0.10, 'Breadth of non-damage tools; a cheaper measure, weighted lower'),
     ('w_typing', 0.11, "Resistance profile against THIS boss's actual attacking types"),
     ('w_niche', 0.10, 'Best archetype fit, so a good specialist is not flattened by averaging')]
for i, (n, v, why) in enumerate(W, 3):
    aw.cell(i, 1, n).font = BLACK
    c = aw.cell(i, 2, v); c.font = BLUE; c.fill = YELL; c.number_format = '0.00'
    aw.cell(i, 3, why).font = Font(name=F, size=10)
aw.cell(10, 1, 'SUM (must be 1.00)').font = BOLD
aw.cell(10, 2, '=SUM(B3:B9)').font = BLACK
aw.cell(10, 2).number_format = '0.00'

aw.cell(11, 1, 'w_sustain (defensive roles)').font = BLACK
_c = aw.cell(11, 2, 0.24); _c.font = BLUE; _c.fill = YELL; _c.number_format = '0.00'
aw.cell(11, 3, 'Walls, clerics, hazard setters and removers, pivots, tanks, trappers, '
               'status spreaders and screen setters use this instead of B5').font = Font(name=F, size=10)
aw.cell(12, 1, 'w_speed (defensive roles)').font = BLACK
_c = aw.cell(12, 2, 0.04); _c.font = BLUE; _c.fill = YELL; _c.number_format = '0.00'
aw.cell(12, 3, 'A wall paying 0.12 on a stat it does not want was a 12-point penalty for '
               'doing its job').font = Font(name=F, size=10)
aw.cell(13, 1, 'defensive SUM (must be 1.00)').font = BOLD
aw.cell(13, 2, '=B3+B4+B11+B12+B7+B8+B9').font = BLACK
aw.cell(13, 2).number_format = '0.00'

aw.cell(15, 1, 'TIER THRESHOLDS on mean VORP (ceiling)').font = BOLD
for j, h in enumerate(['Tier', 'Min VORP', 'Note'], 1):
    c = aw.cell(16, j, h); c.font = HDR; c.fill = HFILL
TH = [('S+', 0.2740), ('S', 0.2085), ('A', 0.1538), ('B', 0.0991), ('C', 0.0555), ('D', -0.0102), ('F', -99)]
for i, (t, lo) in enumerate(TH, 17):
    aw.cell(i, 1, t).font = BLACK
    c = aw.cell(i, 2, lo); c.font = BLUE; c.fill = YELL; c.number_format = '0.00'
aw.cell(24, 3, 'Not curve-forced. If the hack genuinely has many S-tier options, '
               'that is the finding.').font = Font(name=F, size=10, italic=True)

aw.cell(26, 1, 'INVESTMENT COST COMPONENTS').font = BOLD
for j, h in enumerate(['Component', 'Charge', 'Why'], 1):
    c = aw.cell(27, j, h); c.font = HDR; c.fill = HFILL
IC = [('Non-neutral nature', 1.0, 'Mint or breeding needed'),
      ('Held item required', 1.0, 'Ceiling uses an item Phase 2 gates'),
      ('Ability differs from common', 1.0, 'Ability Capsule / Patch'),
      ('Species needs an evolution item', 1.0, 'From availability_basis'),
      ('Per gated TM used (cap 2.0)', 0.5, 'TM competition and cost'),
      ('Any egg move used', 2.0, 'RUN_CONFIG says breeding is NOT used in this run, '
                                 'so an egg-move ceiling is out of reach, not free'),
      ('Floor', 0.5, 'So ROI never divides by zero')]
for i, (n, v, why) in enumerate(IC, 28):
    aw.cell(i, 1, n).font = BLACK
    c = aw.cell(i, 2, v); c.font = BLUE; c.fill = YELL; c.number_format = '0.0'
    aw.cell(i, 3, why).font = Font(name=F, size=10)
aw.cell(36, 1, 'EXCLUDED from cost').font = BOLD
aw.cell(36, 3, 'EV-training time and breeding, per RUN_CONFIG: the eight unconfirmable '
               'fields were dropped in Phase 1. ROI here is therefore NOT comparable to a '
               'version of this analysis with a full cost model.').font = Font(name=F, size=10, italic=True)

# ======================== CHECKPOINTS =================================
cw = wb.create_sheet('Checkpoints')
cols = ['checkpoint', 'checkpoint_label', 'level_cap', 'cap_basis', 'roster_basis',
        'n_anchors', 'n_roster_slots', 'pool_size', 'tms_available', 'items_available',
        'anchor_trainers']
for j, h in enumerate(cols, 1):
    c = cw.cell(1, j, h); c.font = HDR; c.fill = HFILL
for i, r in enumerate(CPT[cols].itertuples(index=False), 2):
    for j, v in enumerate(r, 1):
        cell = cw.cell(i, j, v)
        cell.font = BLUE if cols[j-1] == 'level_cap' else BLACK
        if cols[j-1] == 'cap_basis' and v == 'interpolated':
            cell.fill = YELL
        if cols[j-1] == 'level_cap' and CPT.iloc[i-2].cap_basis == 'interpolated':
            cell.fill = YELL
# monotonicity check as a live formula
cw.cell(1, 13, 'cap > previous?').font = HDR; cw.cell(1, 13).fill = HFILL
for i in range(3, 21):
    cw.cell(i, 13, f'=IF(C{i}>C{i-1},"ok","NOT MONOTONIC")').font = BLACK
cw.cell(22, 1, 'All 19 caps are MEASURED from Boss_Battles.xlsx!Main. The anchor_trainers column '
               'names the fight each checkpoint is scored against; cp09 Pre-Trick House has no '
               'single boss, so its threat pool is the 46 Trick House sheet slots.'
       ).font = Font(name=F, size=10, italic=True)
for j, w in zip('ABCDEFGHIJKM', [11, 26, 10, 13, 30, 9, 15, 10, 15, 16, 60, 16]):
    cw.column_dimensions[j].width = w

# ===================== VALUATION (the work) ===========================
vw = wb.create_sheet('Valuation')
show = ['species_form', 'bst', 'tier', 'primary_niche', 'earliest_checkpoint',
        'niche_stat_S', 'niche_tool_T', 'niche_type_Y', 'primary_niche_fit',
        'value_floor_mean', 'value_mean', 'investment_cost',
        'vorp_mean', 'roi_mean', 'n_checkpoints_obtainable', 'cant_miss', 'confidence']
D = VAL[show].copy().reset_index(drop=True)
# per-axis means, kept visible so the composite is auditable
ax = V.groupby('species').agg(off=('offense_ceiling', 'mean'), dfn=('defense_ceiling', 'mean'),
                              spd=('speed_neutral_ceiling', 'mean'), utl=('utility_ceiling', 'mean'),
                              typ=('typing', 'mean'), rep=('replacement_level', 'mean'), sus=('kit_scaled_ceiling', 'mean'),
                              nfit=('primary_niche_fit', 'mean')).round(4)
D = D.join(ax, on='species_form')
# The composite is a mean of per-checkpoint composites, and mean(sum of terms)
# equals sum(mean of terms) only if EVERY term is the checkpoint mean. Column I
# holds the LAST checkpoint's niche fit (useful context), so the formula must
# read the checkpoint MEAN instead, or it lands systematically low -- it did,
# by ~0.03, which was enough to move Orbeetle-Mega from S+ to S.
D['niche_fit_mean'] = D.pop('nfit')
MODAL = dict(zip(VAL.species_form, VAL.modal_niche))
DEFR = {'physical_wall', 'special_wall', 'mixed_wall', 'bulky_pivot', 'tank', 'cleric',
        'hazard_setter', 'hazard_remover', 'status_spreader', 'screen_setter', 'trapper'}

headers = ['Species form', 'BST', 'Tier (formula)', 'Primary niche', 'First cp',
           'Stat S', 'Tool T', 'Type Y', 'Niche fit',
           'Floor value', 'Ceiling value', 'Invest cost',
           'VORP (formula)', 'ROI (formula)', '# cps', "Can't miss", 'Confidence',
           'Offense', 'Defense', 'Speed (ceiling)', 'Utility', 'Typing', 'Replacement',
           'Sustain (kit)', 'Niche fit (mean)', 'Composite (formula)', 'Def role?']
for j, h in enumerate(headers, 1):
    c = vw.cell(1, j, h); c.font = HDR; c.fill = HFILL
    c.alignment = Alignment(wrap_text=True, vertical='center')
vw.freeze_panes = 'B2'

for i, r in enumerate(D.itertuples(index=False), 2):
    vals = list(r)
    for j, v in enumerate(vals, 1):
        cell = vw.cell(i, j, v if not (isinstance(v, float) and np.isnan(v)) else None)
        cell.font = BLACK
        if headers[j-1] in ('Floor value', 'Ceiling value', 'Niche fit', 'Stat S', 'Tool T',
                            'Type Y', 'Offense', 'Defense', 'Speed', 'Utility', 'Typing',
                            'Replacement'):
            cell.number_format = '0.000'
    # live formulas: composite from the Assumptions weights, then VORP, ROI, tier
    # role-aware: defensive roles swap in the B11/B12 sustain and speed weights
    vw.cell(i, 27, 1 if MODAL.get(D.species_form.iloc[i - 2]) in DEFR else 0).font = BLACK
    vw.cell(i, 26, f'=Assumptions!$B$3*R{i}+Assumptions!$B$4*S{i}'
                   f'+IF(AA{i}=1,Assumptions!$B$11,Assumptions!$B$5)*X{i}'
                   f'+IF(AA{i}=1,Assumptions!$B$12,Assumptions!$B$6)*T{i}'
                   f'+Assumptions!$B$7*U{i}+Assumptions!$B$8*V{i}'
                   f'+Assumptions!$B$9*Y{i}').font = GREEN
    vw.cell(i, 26).number_format = '0.000'
    vw.cell(i, 13, f'=Z{i}-W{i}').font = BLACK
    vw.cell(i, 13).number_format = '0.000'
    vw.cell(i, 14, f'=IFERROR((Z{i}-J{i})/L{i},"")').font = BLACK
    vw.cell(i, 14).number_format = '0.000'
    vw.cell(i, 3, f'=IF(M{i}>=Assumptions!$B$17,"S+",IF(M{i}>=Assumptions!$B$18,"S",'
                  f'IF(M{i}>=Assumptions!$B$19,"A",IF(M{i}>=Assumptions!$B$20,"B",'
                  f'IF(M{i}>=Assumptions!$B$21,"C",IF(M{i}>=Assumptions!$B$22,"D","F"))))))'
            ).font = GREEN
for j, w in zip(range(1, 28), [26, 6, 9, 22, 8, 8, 8, 8, 9, 10, 11, 10, 12, 9, 7, 10, 12,
                               9, 9, 8, 8, 8, 11, 11, 11, 12, 9]):
    vw.column_dimensions[get_column_letter(j)].width = w

# ===================== VALIDATION =====================================
sw = wb.create_sheet('Validation')
for j, h in enumerate(['phase', 'check', 'result', 'severity', 'affected', 'effect'], 1):
    c = sw.cell(1, j, h); c.font = HDR; c.fill = HFILL
for i, r in enumerate(VALID.itertuples(index=False), 2):
    for j, v in enumerate(r, 1):
        c = sw.cell(i, j, v); c.font = BLACK
        c.alignment = Alignment(wrap_text=True, vertical='top')
        if j == 4 and v in ('Material', 'Unresolved'):
            c.fill = YELL
for j, w in zip('ABCDEF', [7, 52, 46, 12, 10, 62]):
    sw.column_dimensions[j].width = w

# ===================== INTEGRITY LOG ==================================
iw = wb.create_sheet('Integrity')
iw.column_dimensions['A'].width = 4
iw.column_dimensions['B'].width = 130
LOGL = [
 'DATA INTEGRITY LOG — Phase 4',
 '',
 'CORRECTIONS TO MY OWN WORK (each would have silently poisoned results):',
 "1. imperium_layer.json exports the Defense base stat as 'df' (the fork's internal compressed "
 "key). The public Specie interface requires 'def'. Unremapped, baseStats.def is undefined and "
 "every physical calc returns NaN rather than erroring. Remapped with a guard that throws.",
 "2. layer['items'] is an ALPHABETICALLY SORTED name list, not an id-indexed array. Using "
 "items[param] for EVO_ITEM resolved id 211 to 'Fire Gem' instead of 'Fire Stone'. The "
 "authoritative source is 02_world.internal_id.",
 "3. 02_world.internal_id is POLYMORPHIC BY entity_type: TM and HM rows store the MOVE id, not "
 "an item id. 28 ids collide with real items — 211 is Fire Stone as an item and Steel Wing as a "
 "TM. TM/HM rows excluded from the item map; all nine evolution stones then resolve correctly.",
 "4. Species were joined layer<->TSV on species_name, which collapsed every alternate form onto "
 "its base ('Venusaur' base and 'Venusaur' Mega are two rows with one name) and silently dropped "
 "510 of 1,534 movepools. internal_index is an exact 1:1 join across all 1,534 rows; switched.",
 "5. An earlier draft asserted Imperium's move table had no healing field and fell back to a "
 "curated recovery name list. It DOES carry flags.healing (36 moves). Recovery gates three wall "
 "archetypes, so this mattered. The flag also marks DRAIN moves, which heal off damage dealt and "
 "must not qualify a wall; split into 15 reliable recovery and 13 drain.",
 '',
 'FRAMEWORK FAILURE AND REBALANCE:',
 "6. primary_niche_fit vs BST came out r=0.745 (ceiling) / 0.705 (floor), ABOVE the ~0.7 "
 "threshold in archetypes.md — the scoring was a BST proxy. Two fixes: toolkit scoring restricted "
 "to role-relevant tools (breadth is itself a BST proxy), and a stat-SHAPE term blended in so a "
 "480-BST species with 130 Atk reads as more sweeper-shaped than a flat 600. After: r=0.476 / 0.365.",
 '',
 'DATA GAPS CARRIED, NOT SMOOTHED:',
 "7. RESOLVED. The level-cap ladder is now MEASURED from Boss_Battles.xlsx!Main, supplied after "
 "the first Phase 4 pass. The earlier reconstruction agreed on 18 of 19 caps; cp09 Pre-Trick House "
 "was interpolated at 53 and is actually 56. Checkpoint 19 is labelled Pre-Champion, not Pre-Elite4.",
 "8. WHY THE LADDER WENT MISSING: the Phase 1/2 README declared Boss_Battles.xlsx superseded "
 "because its content appeared verbatim inside General_Documents___Locations.xlsx, on a check of "
 "'content-row counts identical across all 11 boss sheets'. The workbook has TWELVE sheets. The "
 "twelfth is Main, and Main carries the ladder. A supersede check that enumerates only the sheets "
 "it already knows about cannot detect one it does not. The rule this changes: compare SHEET "
 "COUNTS before comparing row counts within sheets.",
 "9. cp09 Pre-TrickHouse has no single boss fight, but the checkpoint IS the Trick House, so its "
 "46 Trick House sheet slots are the correct threat pool rather than a stand-in. Downgraded from "
 "Unresolved to Material.",
 "10. Phase 3 parsed 678 roster slots against 695 'Ability:' lines in the source. The 17-row gap "
 "is exactly the 17 megas' post-transformation ability lines, which Phase 3 correctly folded into "
 "mega_ability (per-sheet mega counts 3/1/3/1/5/4 match the per-sheet gaps exactly). Nothing is "
 "missing; RUN_CONFIG's '695 roster slots' overcounts real Pokemon by 17.",
 "11. Battle_Rewards.xlsx re-checked: all 17 rows already present in 02_world with matching "
 "locations, so that supersede claim held. Its trainer column was NOT in 02_world and has been "
 "added as reward_trainer, upgrading those rows' gate_basis from 'location reachability' (a lower "
 "bound) to 'defeat <trainer>' (a sufficient condition).",
 "cp02 uses the 30-slot DAWN:16 block, which the Phase 3 parse merged from several rival "
 "fights. Roster breadth at cp02 is therefore overstated.",
 "12. 24 evolution methods remain undecoded integers (from Phase 1). Species reached through them "
 "are dated to their pre-evolution and marked Unresolved.",
 "13. 223 of 1,534 forms have no derivable acquisition route (192 already Phase-1 unreachable, "
 "17 gigantamax with no G-Max item or mechanic in any source, 9 megas whose stone did not match, "
 "5 others). They are excluded from every pool rather than given a guessed date.",
 "14. 7 boss-roster held items are absent from Imperium's own itemData (Heavy-Duty Boots, Weakness "
 "Policy, Terrain Extender, Hearthflame/Wellspring Mask, Flapplite, type Memories) and 1 ability "
 "(As One (Spectrier)). Flagged Unresolved on their slot, never substituted.",
 "15. weightkg is 50 for every species — no Imperium source carries weight. Low Kick, Heavy Slam "
 "and Grass Knot are Unresolved, not merely Inferred.",
 "16. Move priority is absent from Imperium's move table and defaults to 0, which affects the "
 "Revenge Killer gate specifically.",
 '',
 'NOT A BUG — verified against the data:',
 "17. Flash is a 60 BP Electric SPECIAL attack in Imperium, not the vanilla accuracy-drop. Move "
 "ids check out against vanilla (Growl 45, Recover 105, Swords Dance 14), so this is a real hack "
 "change, not a parse offset.",
 "18. Volbeat's 145-171% Bug Buzz on Lokix looked wrong and is correct: Imperium's Bug->Dark is "
 "2.0 and the hand formula reproduces 132-156 exactly.",
]
for i, t in enumerate(LOGL, 1):
    c = iw.cell(i, 2, t)
    c.font = BOLD if t.isupper() or t.endswith(':') or i == 1 else Font(name=F, size=10)
    c.alignment = Alignment(wrap_text=True, vertical='top')

# ===================== CHARTS =========================================
gw = wb.create_sheet('Charts')
tiers = VAL.tier.value_counts().reindex(['S+', 'S', 'A', 'B', 'C', 'D', 'F']).fillna(0)
gw.cell(1, 1, 'Tier').font = BOLD; gw.cell(1, 2, 'Species').font = BOLD
for i, (t, n) in enumerate(tiers.items(), 2):
    gw.cell(i, 1, t).font = BLACK; gw.cell(i, 2, int(n)).font = BLACK
ch = BarChart(); ch.title = 'Tier distribution (not curve-forced)'
ch.y_axis.title = 'species forms'; ch.x_axis.title = 'tier'
ch.add_data(Reference(gw, min_col=2, min_row=1, max_row=8), titles_from_data=True)
ch.set_categories(Reference(gw, min_col=1, min_row=2, max_row=8))
ch.height, ch.width = 8, 14
gw.add_chart(ch, 'E2')

pool = CPT[['checkpoint', 'pool_size', 'tms_available', 'level_cap']]
gw.cell(11, 1, 'Checkpoint').font = BOLD
gw.cell(11, 2, 'Obtainable pool').font = BOLD
gw.cell(11, 3, 'TMs available').font = BOLD
gw.cell(11, 4, 'Level cap').font = BOLD
for i, r in enumerate(pool.itertuples(index=False), 12):
    for j, v in enumerate(r, 1):
        gw.cell(i, j, int(v)).font = BLACK
lc = LineChart(); lc.title = 'What the player can actually field, by checkpoint'
lc.y_axis.title = 'count'; lc.x_axis.title = 'checkpoint'
lc.add_data(Reference(gw, min_col=2, max_col=4, min_row=11, max_row=30), titles_from_data=True)
lc.set_categories(Reference(gw, min_col=1, min_row=12, max_row=30))
lc.height, lc.width = 8, 14
gw.add_chart(lc, 'E20')

# BST vs niche fit — the evidence the framework is not a BST proxy
samp = VAL[['bst', 'primary_niche_fit']].dropna().sample(min(400, len(VAL)), random_state=1)
gw.cell(34, 1, 'BST').font = BOLD; gw.cell(34, 2, 'primary_niche_fit').font = BOLD
for i, r in enumerate(samp.itertuples(index=False), 35):
    gw.cell(i, 1, int(r.bst)).font = BLACK
    gw.cell(i, 2, float(r.primary_niche_fit)).font = BLACK
sc = ScatterChart(); sc.title = 'BST vs primary_niche_fit (r = 0.476) — no BST proxy'
sc.x_axis.title = 'BST'; sc.y_axis.title = 'niche fit'; sc.style = 13
s = Series(Reference(gw, min_col=2, min_row=35, max_row=34 + len(samp)),
           Reference(gw, min_col=1, min_row=35, max_row=34 + len(samp)), title='species')
s.marker.symbol = 'circle'; s.graphicalProperties.line.noFill = True
sc.series.append(s)
sc.height, sc.width = 9, 14
gw.add_chart(sc, 'E38')

wb.save(OUT)
print('wrote', OUT)
