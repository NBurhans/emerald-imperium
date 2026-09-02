"""Audit the Phase 3 parse against Boss_Battles.xlsx directly.

The previous README declared this file superseded on the grounds that its
content appeared verbatim inside General_Documents___Locations.xlsx, with
'content-row counts identical across all 11 boss sheets'. That check counted
ELEVEN sheets. This workbook has TWELVE. The twelfth is `Main`, and `Main` is
the sheet carrying the level-cap ladder -- the one thing later phases could
not derive. A supersede check that compares only the sheets it already knows
about cannot detect a sheet it does not.

So this re-audits everything rather than re-asserting the same claim:
  1. the cap ladder, now Measured
  2. roster slot counts per sheet, source vs 03_trainer_slots.tsv
  3. the RUN_CONFIG anchor counts (695 Ability: lines, 45 IV lines, 9 '0 Spe')
"""
from openpyxl import load_workbook
import pandas as pd, re, json

U = '/mnt/user-data/uploads/'
P = '/home/claude/p4/'
wb = load_workbook(U + 'Boss_Battles.xlsx', read_only=True)
print('sheets in Boss_Battles.xlsx:', len(wb.sheetnames))
print(' ', wb.sheetnames)

# ---------- 1. cap ladder, Measured ----------
ws = wb['Main']
caps = []
seen_header = False
for row in ws.iter_rows(values_only=True):
    vals = [c for c in row if c is not None and str(c).strip() != '']
    if not vals:
        continue
    if str(vals[0]).strip() == 'Level Caps':
        seen_header = True
        continue
    if seen_header:
        if len(vals) >= 2 and isinstance(vals[1], (int, float)):
            caps.append((str(vals[0]).strip(), int(vals[1])))
        elif caps:
            break
print(f'\n--- LEVEL CAP LADDER (Measured, Boss_Battles!Main) --- {len(caps)} entries')
mine = [15, 20, 25, 30, 32, 34, 44, 47, 53, 59, 64, 68, 71, 74, 76, 80, 82, 84, 85]
print(f'{"#":>3} {"source label":34} {"cap":>4} {"my reconstruction":>18} {"delta":>6}')
deltas = []
for i, (lbl, cap) in enumerate(caps):
    d = cap - mine[i] if i < len(mine) else None
    deltas.append(d)
    flag = '' if d == 0 else '   <-- CORRECTED'
    print(f'{i+1:3d} {lbl:34} {cap:4d} {mine[i] if i < len(mine) else "-":>18} {d:+6d}{flag}')
print(f'exact: {sum(1 for d in deltas if d == 0)}/{len(caps)}')

# ---------- 2. roster slots, source vs parse ----------
SHEETS = [s for s in wb.sheetnames if s != 'Main']
src = {}
iv_lines, zero_spe, ev_lines, hp_moves = 0, 0, 0, 0
mons_per_sheet = {}
for sh in SHEETS:
    w = wb[sh]
    n_ab = 0
    for row in w.iter_rows(values_only=True):
        for c in row:
            if c is None:
                continue
            s = str(c).strip()
            if s.startswith('Ability:'):
                n_ab += 1
            elif s.startswith('IVs:'):
                iv_lines += 1
                if re.search(r'\b0 Spe\b', s):
                    zero_spe += 1
            elif s.startswith('EVs:'):
                ev_lines += 1
            elif s.startswith('- Hidden Power'):
                hp_moves += 1
    src[sh] = n_ab
    mons_per_sheet[sh] = n_ab

SL = pd.read_csv(P + '03_trainer_slots.tsv', sep='\t', dtype=str)
parsed = SL.sheet.value_counts().to_dict()

print('\n--- ROSTER SLOTS: source "Ability:" lines vs 03_trainer_slots.tsv ---')
print(f'{"sheet":20} {"source":>8} {"parsed":>8} {"delta":>7}')
tot_s = tot_p = 0
gaps = []
for sh in SHEETS:
    s, p = src[sh], parsed.get(sh, 0)
    tot_s += s; tot_p += p
    if s != p:
        gaps.append((sh, s, p))
    print(f'{sh:20} {s:8d} {p:8d} {p-s:+7d}{"   <-- GAP" if s != p else ""}')
print(f'{"TOTAL":20} {tot_s:8d} {tot_p:8d} {tot_p-tot_s:+7d}')

print('\n--- RUN_CONFIG anchor counts, re-measured against this file ---')
print(f'{"claim":46} {"RUN_CONFIG":>11} {"measured":>9}')
for lbl, claimed, got in [
        ('Total roster slots (Ability: lines)', 695, tot_s),
        ('Slots annotated with an IVs: line', 45, iv_lines),
        ("Slots annotated 'IVs: 0 Spe'", 9, zero_spe),
        ('Slots carrying EV spreads', 576, ev_lines),
        ('Hidden Power moves on rosters', 39, hp_moves)]:
    print(f'{lbl:46} {claimed:11d} {got:9d}{"" if claimed == got else "   <-- DIFFERS"}')

json.dump({'caps': caps, 'src_slots': src, 'iv_lines': iv_lines,
           'zero_spe': zero_spe, 'ev_lines': ev_lines, 'hp_moves': hp_moves},
          open(P + 'boss_audit.json', 'w'), indent=1)
print('\nwrote boss_audit.json')
