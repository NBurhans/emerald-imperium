"""Add the reward TRAINER to 02_world's held_item rows.

Battle_Rewards.xlsx carries a column 02_world does not: which trainer or
battle yields each held item. 02_world already has all 17 rows with matching
locations, so nothing was lost when this file was set aside -- but the trainer
attribution is real information that was dropped, and it upgrades those rows'
gate_basis from 'location reachability' (an explicit LOWER bound: reaching the
place is necessary but the source did not confirm it is sufficient) to
'defeat <trainer>', which is a sufficient condition.

The gate CHECKPOINT is not changed here. Naming the trainer does not by itself
tell us when that trainer is fightable; that would need the trainer pinned to
a checkpoint, and most of these are Sinnoh leaders and mini-bosses whose
fights scale to the player. The added value is that the basis is now honest
about what kind of condition it is.
"""
from openpyxl import load_workbook
import pandas as pd, re

U = '/mnt/user-data/uploads/'
P = '/home/claude/p4/'

wb = load_workbook(U + 'Battle_Rewards.xlsx', read_only=True)
ws = wb['Battle Rewards']
rows = list(ws.iter_rows(values_only=True))
hdr = [str(c).strip() if c else '' for c in rows[0]]
recs = []
for r in rows[1:]:
    if not r or r[0] is None:
        continue
    recs.append(dict(zip(hdr, [str(c).strip() if c is not None else '' for c in r])))
BR = pd.DataFrame(recs)
print(f'Battle_Rewards rows parsed: {len(BR)}')


def n(s):
    return re.sub(r'[^a-z0-9]', '', str(s).lower())


W = pd.read_csv(P + '02_world.tsv', sep='\t', dtype=str)
if 'reward_trainer' not in W.columns:
    W['reward_trainer'] = 'NA'

# match on (item name, location) so the two Leftovers rows -- Route 117 from
# Fantina and Route 121 from Byron -- do not collapse onto each other.
key = {(n(r['Reward Item']), n(r['Location'])): r['Trainer/Battle'] for _, r in BR.iterrows()}
matched, unmatched = 0, []
for i, r in W.iterrows():
    if r.entity_type != 'held_item':
        continue
    k = (n(r.entity_name), n(r.location))
    if k in key:
        W.at[i, 'reward_trainer'] = key[k]
        W.at[i, 'gate_basis'] = f'defeat {key[k]}'
        matched += 1
    else:
        unmatched.append((r.entity_name, r.location))

print(f'held_item rows enriched: {matched} of {(W.entity_type == "held_item").sum()}')
if unmatched:
    print('UNMATCHED (kept as-is, not guessed):', unmatched)

# every Battle_Rewards row should have found a home
srckeys = set(key)
hit = {(n(r.entity_name), n(r.location)) for _, r in W[W.entity_type == 'held_item'].iterrows()}
orphan = srckeys - hit
print(f'Battle_Rewards rows with no 02_world row: {len(orphan)} {sorted(orphan) if orphan else ""}')

W.to_csv(P + '02_world.tsv', sep='\t', index=False)
print('\nheld_item rows after enrichment:')
print(W[W.entity_type == 'held_item'][
    ['entity_name', 'location', 'reward_trainer', 'earliest_gate', 'gate_basis']
].to_string(index=False))
